"""
俸給表データ（data/salary-tables-r8.json 等）の生成スクリプト。

公式の俸給表xlsx（人事院公表の別表第一〜第十等に相当する各シート、列D以降が
各級、行が号俸というレイアウト）を読み込み、俸給表JSONを再生成する。

使い方:
    python3 scripts/extract-salary-tables.py path/to/official-salary-tables.xlsx \
        [output.json] [effectiveDate] [source note]

前提とするシート構成:
- 「職務の級」を含むヘッダーセルの右側に級番号が並ぶ「級構成」シート
  （行政職(一)・専門行政職・税務職・公安職(一)(二)・海事職(一)(二)・
   教育職(一)(二)・研究職・医療職(一)(二)(三)・福祉職・専門スタッフ職）
- 号俸と俸給月額のみの「フラット構成」シート（指定職）
- 複数のフラット表が縦に並ぶ特殊シート（任期付研究員・特定任期付）

ユーザー指示により行政職俸給表(二)は対象外（GRADED_SHEETSに含めていない）。

各俸給表のシート名は候補を複数持たせてあり（例: 「行政職（一）」「行（一）」）、
人事院勧告資料ごとの略記ゆれに対応する。新しいレイアウトのファイルを渡す場合は
GRADED_SHEETS 等の定義を書き換えること。
"""

import json
import sys

import openpyxl

GRADED_SHEETS = {
    "administrative_1": ("行政職俸給表(一)", ["行政職（一）", "行（一）"]),
    "specialized_administrative": ("専門行政職俸給表", ["専門行政職", "専門行政"]),
    "tax": ("税務職俸給表", ["税務職", "税務"]),
    "public_safety_1": ("公安職俸給表(一)", ["公安職（一）", "公（一）"]),
    "public_safety_2": ("公安職俸給表(二)", ["公安職（二）", "公（二）"]),
    "maritime_1": ("海事職俸給表(一)", ["海事職（一）", "海（一）"]),
    "maritime_2": ("海事職俸給表(二)", ["海事職（二）", "海（二）"]),
    "education_1": ("教育職俸給表(一)", ["教育職（一）", "教（一）"]),
    "education_2": ("教育職俸給表(二)", ["教育職（二）", "教（二）"]),
    "research": ("研究職俸給表", ["研究職", "研究"]),
    "medical_1": ("医療職俸給表(一)", ["医療職（一）", "医（一）"]),
    "medical_2": ("医療職俸給表(二)", ["医療職（二）", "医（二）"]),
    "medical_3": ("医療職俸給表(三)", ["医療職（三）", "医（三）"]),
    "welfare": ("福祉職俸給表", ["福祉職", "福祉"]),
    "specialist_staff": ("専門スタッフ職俸給表", ["専門スタッフ職", "専スタ"]),
}

DESIGNATED_SHEET_NAMES = ["指定職", "指定"]
FIXED_TERM_SHEET_NAMES = ["任期付研究員・特定任期付"]


def find_sheet(wb, candidates):
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]
    return None


def parse_graded_sheet(ws):
    header_row = None
    grade_col_start = None
    for r in range(1, 8):
        for c in range(1, 6):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and "職務" in v and "級" in v:
                header_row = r
                grade_col_start = c + 1
                break
        if header_row:
            break
    if not header_row:
        return None

    grades = {}
    grade_cols = {}
    c = grade_col_start
    while True:
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, int):
            grade_cols[c] = v
            grades[str(v)] = []
            c += 1
        else:
            break

    step_col = grade_col_start - 2

    data_start = None
    for r in range(header_row + 1, header_row + 6):
        if ws.cell(row=r, column=step_col).value == 1:
            data_start = r
            break
    if data_start is None:
        data_start = header_row + 3  # フォールバック（従来レイアウト想定）

    r = data_start
    while r <= ws.max_row:
        stop = False
        for c in range(1, step_col):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                stop = True
                break
        if stop:
            break  # "定年前再任用短時間勤務職員" 区分や備考の開始
        step_val = ws.cell(row=r, column=step_col).value
        row_vals = {gc: ws.cell(row=r, column=gc).value for gc in grade_cols}
        if step_val is None and all(v is None for v in row_vals.values()):
            r += 1
            continue
        if step_val is not None:
            for gc, grade_num in grade_cols.items():
                amt = row_vals[gc]
                if amt is not None:
                    grades[str(grade_num)].append(amt)
        r += 1
        if r - data_start > 400:
            break

    return {k: v for k, v in grades.items() if v}


def parse_flat_table(ws, data_start_row, step_col, val_col, stop_after_blank=1):
    steps = []
    r = data_start_row
    blanks = 0
    while r <= ws.max_row:
        step_val = ws.cell(row=r, column=step_col).value
        amt = ws.cell(row=r, column=val_col).value
        if step_val is None and amt is None:
            blanks += 1
            if blanks > stop_after_blank:
                break
            r += 1
            continue
        blanks = 0
        if amt is not None:
            steps.append(amt)
        r += 1
    return steps


def find_flat_series(ws):
    """シート内の「号俸」ヘッダーを総当たりで探し、各ヘッダーにつながる俸給月額の
    縦系列を1つのフラット表として抽出する。同一シートに複数のフラット表が縦に
    並ぶレイアウト（指定職、任期付研究員・特定任期付）に対応するため、行番号を
    ハードコードせず動的に検出する。"""
    found = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if not (isinstance(v, str) and v.strip() == "号俸"):
                continue
            step_col = c
            val_col = None
            for vc in range(c + 1, min(c + 6, ws.max_column + 1)):
                hv = ws.cell(row=r, column=vc).value
                if isinstance(hv, str):
                    compact = hv.replace("　", "").replace(" ", "").replace("\n", "")
                    if "俸給" in compact and "月額" in compact:
                        val_col = vc
                        break
            if val_col is None:
                continue
            steps = parse_flat_table(ws, data_start_row=r + 2, step_col=step_col, val_col=val_col, stop_after_blank=0)
            if steps:
                found.append((r, steps))
    found.sort(key=lambda x: x[0])
    return [steps for _, steps in found]


def main():
    if len(sys.argv) < 2:
        print(f"usage: {sys.argv[0]} <official-salary-tables.xlsx> [output.json] [effectiveDate] [source note]")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else "data/salary-tables-r8.json"

    from datetime import date

    effective_date = sys.argv[3] if len(sys.argv) > 3 else date.today().isoformat()
    source_note = (
        sys.argv[4]
        if len(sys.argv) > 4
        else "ユーザー提供の公式俸給表データ（一般職の職員の給与に関する法律 別表に基づく）"
    )

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    order_keys = []
    tables = {}

    for key, (label, sheet_names) in GRADED_SHEETS.items():
        ws = find_sheet(wb, sheet_names)
        if ws is None:
            print(f"skip (sheet not found): {key} (候補: {sheet_names})")
            continue
        grades = parse_graded_sheet(ws)
        if not grades:
            print(f"FAILED TO PARSE: {key}")
            continue
        tables[key] = {"label": label, "type": "graded", "grades": grades}
        order_keys.append(key)
        print(key, "->", {g: len(v) for g, v in grades.items()})

    ws = find_sheet(wb, DESIGNATED_SHEET_NAMES)
    if ws is not None:
        series = find_flat_series(ws)
        if series:
            tables["designated"] = {"label": "指定職俸給表", "type": "flat", "steps": series[0]}
            order_keys.append("designated")
            print("designated ->", series[0])

    ws = find_sheet(wb, FIXED_TERM_SHEET_NAMES)
    if ws is not None:
        series = find_flat_series(ws)
        fixed_term_defs = [
            ("fixed_term_researcher_type1", "任期付研究員（第一号）俸給表（号俸のみ）"),
            ("fixed_term_researcher_type2", "任期付研究員（第二号）俸給表（号俸のみ）"),
            ("fixed_term_staff_specified", "特定任期付職員俸給表（号俸のみ）"),
        ]
        if len(series) != len(fixed_term_defs):
            print(f"WARNING: expected {len(fixed_term_defs)} flat series in 任期付研究員・特定任期付, found {len(series)}")
        for (key, label), steps in zip(fixed_term_defs, series):
            tables[key] = {"label": label, "type": "flat", "steps": steps}
            order_keys.append(key)
            print(key, "->", steps)

    out = {
        "source": source_note,
        "note": f"提供時点（{date.today().isoformat()}）で有効な俸給表として掲載。実際の施行日は前年度分の可能性があるため、最新の人事院公表資料と照合してください。",
        "effectiveDate": effective_date,
        "order": order_keys,
        "tables": tables,
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=1)

    print(f"wrote {output_path} with", len(tables), "tables")


if __name__ == "__main__":
    main()
